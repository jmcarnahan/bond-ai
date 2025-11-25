"""
Unit tests for XLSM to XLSX conversion functionality.

This test suite verifies that XLSM (macro-enabled Excel) files are correctly
converted to XLSX format by removing macros while preserving data.
"""

import pytest
import io
import openpyxl
from bondable.bond.providers.files import convert_xlsm_to_xlsx


class TestXLSMConversion:
    """Test the convert_xlsm_to_xlsx function"""

    @pytest.fixture
    def sample_xlsm_bytes(self):
        """Create a sample XLSM file with data (no actual macros for testing)"""
        # Create a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Data"

        # Add some sample data
        sheet['A1'] = "Name"
        sheet['B1'] = "Value"
        sheet['A2'] = "Test 1"
        sheet['B2'] = 100
        sheet['A3'] = "Test 2"
        sheet['B3'] = 200
        sheet['A4'] = "Test 3"
        sheet['B4'] = 300

        # Add a formula
        sheet['B5'] = "=SUM(B2:B4)"

        # Save as BytesIO (would be XLSM format in real file)
        file_bytes = io.BytesIO()
        workbook.save(file_bytes)
        file_bytes.seek(0)

        return file_bytes

    def test_convert_valid_xlsm(self, sample_xlsm_bytes):
        """Test converting a valid XLSM file to XLSX"""
        original_filename = "test_data.xlsm"

        # Convert
        converted_bytes, new_mime_type, new_filename = convert_xlsm_to_xlsx(
            sample_xlsm_bytes, original_filename
        )

        # Verify return types
        assert isinstance(converted_bytes, io.BytesIO)
        assert isinstance(new_mime_type, str)
        assert isinstance(new_filename, str)

        # Verify MIME type changed
        assert new_mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Verify filename extension changed
        assert new_filename == "test_data.xlsx"

        # Verify converted file is valid Excel
        converted_bytes.seek(0)
        workbook = openpyxl.load_workbook(converted_bytes)
        assert workbook is not None

    def test_data_preservation(self, sample_xlsm_bytes):
        """Test that data is preserved after conversion"""
        original_filename = "data.xlsm"

        # Get original data
        sample_xlsm_bytes.seek(0)
        original_workbook = openpyxl.load_workbook(sample_xlsm_bytes)
        original_sheet = original_workbook.active
        original_value = original_sheet['B2'].value

        # Reset for conversion
        sample_xlsm_bytes.seek(0)

        # Convert
        converted_bytes, _, _ = convert_xlsm_to_xlsx(sample_xlsm_bytes, original_filename)

        # Load converted file
        converted_bytes.seek(0)
        converted_workbook = openpyxl.load_workbook(converted_bytes)
        converted_sheet = converted_workbook.active

        # Verify data preserved
        assert converted_sheet['A1'].value == "Name"
        assert converted_sheet['B1'].value == "Value"
        assert converted_sheet['B2'].value == original_value
        assert converted_sheet['A2'].value == "Test 1"

    def test_extension_change_case_insensitive(self):
        """Test that extension change works for different cases"""
        test_cases = [
            ("file.xlsm", "file.xlsx"),
            ("FILE.XLSM", "FILE.xlsx"),
            ("MixedCase.XlSm", "MixedCase.xlsx"),
            ("multiple.dots.xlsm", "multiple.dots.xlsx"),
        ]

        # Create a minimal workbook
        workbook = openpyxl.Workbook()
        workbook.active['A1'] = "test"
        file_bytes = io.BytesIO()
        workbook.save(file_bytes)

        for original, expected in test_cases:
            file_bytes.seek(0)
            _, _, new_filename = convert_xlsm_to_xlsx(file_bytes, original)
            assert new_filename == expected, f"Failed for {original}"

    def test_filename_without_xlsm_extension(self):
        """Test handling of filename without .xlsm extension"""
        # Create a minimal workbook
        workbook = openpyxl.Workbook()
        workbook.active['A1'] = "test"
        file_bytes = io.BytesIO()
        workbook.save(file_bytes)
        file_bytes.seek(0)

        original_filename = "somefile.txt"
        _, _, new_filename = convert_xlsm_to_xlsx(file_bytes, original_filename)

        # Should not change if doesn't end with .xlsm
        assert new_filename == original_filename

    def test_corrupt_file_handling(self):
        """Test that corrupt files raise appropriate exceptions"""
        corrupt_bytes = io.BytesIO(b"This is not an Excel file")

        with pytest.raises(Exception):
            convert_xlsm_to_xlsx(corrupt_bytes, "corrupt.xlsm")

    def test_empty_file_handling(self):
        """Test that empty files raise appropriate exceptions"""
        empty_bytes = io.BytesIO(b"")

        with pytest.raises(Exception):
            convert_xlsm_to_xlsx(empty_bytes, "empty.xlsm")

    def test_mime_type_always_xlsx(self, sample_xlsm_bytes):
        """Test that MIME type is always set to XLSX regardless of input"""
        test_filenames = [
            "test.xlsm",
            "test.XLSM",
            "test.txt",  # Wrong extension
        ]

        for filename in test_filenames:
            sample_xlsm_bytes.seek(0)
            _, mime_type, _ = convert_xlsm_to_xlsx(sample_xlsm_bytes, filename)
            assert mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_workbook_structure_preserved(self, sample_xlsm_bytes):
        """Test that workbook structure (sheets, names) is preserved"""
        # Create a workbook with multiple sheets
        workbook = openpyxl.Workbook()
        workbook.active.title = "Sheet1"
        workbook.create_sheet("Sheet2")
        workbook.create_sheet("Data")

        workbook["Sheet1"]['A1'] = "Sheet 1 Data"
        workbook["Sheet2"]['A1'] = "Sheet 2 Data"
        workbook["Data"]['A1'] = "Data Sheet"

        file_bytes = io.BytesIO()
        workbook.save(file_bytes)
        file_bytes.seek(0)

        # Convert
        converted_bytes, _, _ = convert_xlsm_to_xlsx(file_bytes, "multi.xlsm")

        # Load and verify
        converted_bytes.seek(0)
        converted_workbook = openpyxl.load_workbook(converted_bytes)

        # Verify sheet names
        assert "Sheet1" in converted_workbook.sheetnames
        assert "Sheet2" in converted_workbook.sheetnames
        assert "Data" in converted_workbook.sheetnames

        # Verify data in each sheet
        assert converted_workbook["Sheet1"]['A1'].value == "Sheet 1 Data"
        assert converted_workbook["Sheet2"]['A1'].value == "Sheet 2 Data"
        assert converted_workbook["Data"]['A1'].value == "Data Sheet"


if __name__ == "__main__":
    # Run the tests
    pytest.main([__file__, "-v", "-s"])
